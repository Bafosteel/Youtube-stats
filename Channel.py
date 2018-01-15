import pandas as pd
import requests
import json
from bs4 import BeautifulSoup
import urllib
from tkinter import *
from tkinter.filedialog import askdirectory,askopenfilename
import openpyxl

from os import listdir, path
class Gui(Toplevel):
    def __init__(self, parent, title="Обработка файлов"):
        Toplevel.__init__(self, parent)
        parent.geometry("250x250+100+150")
        if title:
            self.title(title)
        parent.withdraw()
        self.parent = parent
        self.result = None
        dialog = Frame(self)
        self.initial_focus = self.dialog(dialog)
        self.protocol("WM_DELETE_WINDOW", self.on_exit)
        dialog.pack()

    def on_exit(self):
        self.quit()

    def text_3_on(self):
        if self.var_1.get():
            self.text_3["state"] = "normal"
            self.text_3.delete(0, END)
            self.text_3.insert(END, "Name_{{initial_image_name}}")
        else:
            self.text_3["state"] = "disabled"

    def search_folder_for_files(self):
        path_to = askopenfilename()
        print(path_to)
        self.text_1.delete(0, END)
        self.text_1.insert(END, path_to)

    def search_folder_for_new_excel_file(self):
        path_to = askdirectory()
        print(path_to)
        self.text_2.delete(0, END)
        self.text_2.insert(END, path_to)

    def save_to_csv(self,tittle, view_count, comment_count, subscriber_count, video_count,
                                date_of_creation_of_channel, country, privacy_status, keywords, moderate_comments,
                    date_of_upload_latest_video):

        table = pd.DataFrame({'Название канала': tittle, 'Общее количество просмотров': view_count,
                              'Общее количество комментариев': comment_count,
                              'Количество подписчиков': subscriber_count, 'Количество видеороликов': video_count,
                             'Дата создания канала':date_of_creation_of_channel, 'Страна':country,
                              'Доступность канала':privacy_status, 'Ключевые слова': keywords,
                              'Модерация комментариев': moderate_comments,
                              'Дата загрузки последного видеоролика': date_of_upload_latest_video})
        if self.var_1.get():
            table.to_csv(str(self.text_2.get()) + '/' + str(self.text_3.get()) + '.csv', sep=';', index=False)
        else:
            table.to_csv(str(self.text_2.get()) + '/' + 'Данные о каналах' + '.csv', sep=';', index=False)
        import ctypes
        message = 'Готово!'
        ctypes.windll.user32.MessageBoxW(0, message, 'Данные о каналах', 0)
        print('ok')
        return {}

    def get_channel_id(self, channel_name):
        id_channels_list = []
        for name in channel_name:
            try:
                content = requests.get('https://www.googleapis.com/youtube/v3/videos?id=' + str(
                    name) + '&key={API_KEY}&part=snippet')
                data = content.json()
                channel_id = data["items"][0]["snippet"]['channelId']
                print(channel_id)
                id_channels_list.append(channel_id)
            except:
                index = name.find('=')
                content = requests.get('https://www.googleapis.com/youtube/v3/videos?id=' + str(
                    name[index+1:]) + '&key={API_KEY}&part=snippet')
                data = content.json()
                channel_id = data["items"][0]["snippet"]['channelId']
                print(channel_id)
                id_channels_list.append(channel_id)
        return id_channels_list

    def get_channel_name_list(self):
        names = []
        wb = openpyxl.load_workbook(str(self.text_1.get()).replace('/', '\\'))
        sheet = wb.worksheets[0]
        for i in range(1, sheet.max_row):
            if ((sheet.cell(row=i, column=1).value) == None):
                max_row = i - 1
                break
            else:
                max_row = sheet.max_row

        for i in range(1, max_row + 1):
            names.append(sheet.cell(row=i, column=1).value)
        print(names)
        return names

    def start(self):
        channel_name = self.get_channel_name_list()
        id_channels_list = self.get_channel_id(channel_name)
        tittle = []
        view_count = []
        comment_count = []
        subscriber_count = []
        video_count = []
        date_of_creation_of_channel = []
        country = []
        privacy_status = []
        keywords = []
        moderate_comments = []
        date_of_upload_latest_video = []


        for name in id_channels_list:
            content = requests.get('https://www.googleapis.com/youtube/v3/channels?id=' + str(
                name) + '&key={API_KEY}&part=snippet,statistics,status,brandingSettings')
            data = content.json()
            print(" ")
            print(data["items"][0]["snippet"]["title"])
            print(data["items"][0]["statistics"]["viewCount"])
            print(data["items"][0]["statistics"]["commentCount"])
            print(data["items"][0]["statistics"]["subscriberCount"])
            print(data["items"][0]["statistics"]["videoCount"])
            tittle.append(data["items"][0]["snippet"]["title"])
            view_count.append(data["items"][0]["statistics"]["viewCount"])
            comment_count.append(data["items"][0]["statistics"]["commentCount"])
            subscriber_count.append(data["items"][0]["statistics"]["subscriberCount"])
            video_count.append(data["items"][0]["statistics"]["videoCount"])
            date_of_creation_of_channel.append(data["items"][0]["snippet"]["publishedAt"])
            try:
                country.append(data["items"][0]["snippet"]["country"])
            except:
                country.append("None")

            try:
                privacy_status.append(data["items"][0]["status"]["privacyStatus"])
            except:
                privacy_status.append("None")

            try:
                keywords.append(data["items"][0]["brandingSettings"]["channel"]["keywords"])
            except:
                keywords.append("None")

            try:
                moderate_comments.append(data["items"][0]["brandingSettings"]["channel"]["moderateComments"])
            except:
                moderate_comments.append("None")


        for name in id_channels_list:
            response = requests.get('https://www.googleapis.com/youtube/v3/activities?channelId=' + str(
                name) + '&key={API_KEY}&part=snippet')
            last_upload = response.json()
            date_of_upload_latest_video.append(last_upload["items"][0]["snippet"]["publishedAt"])

        return self.save_to_csv(tittle, view_count, comment_count, subscriber_count, video_count,
                                date_of_creation_of_channel, country, privacy_status, keywords, moderate_comments,
                                date_of_upload_latest_video)


    def dialog(self, parent):
        self.parent = parent

        # Created main elements
        self.label_1 = Label(parent, text="Укажите путь, по которому лежит основной Excel файл")
        self.text_1 = Entry(parent, width=50)
        self.but_1 = Button(parent, text="Указать", command=self.search_folder_for_files)

        self.var_1 = IntVar()

        self.label_2 = Label(parent, text="Укажие папку, куда положить получившуюся таблицу")
        self.text_2 = Entry(parent, width=50)
        self.but_2 = Button(parent, text="Указать", command=self.search_folder_for_new_excel_file)

        self.chk_1 = Checkbutton(parent, text="Переименовать файл по маске", variable=self.var_1, command=self.text_3_on)
        self.text_3 = Entry(parent, width=50, state=DISABLED, disabledforeground=parent.cget('bg'))

        self.label_1.pack()
        self.text_1.pack()
        self.but_1.pack()

        self.label_2.pack()
        self.text_2.pack()
        self.but_2.pack()

        self.chk_1.pack()
        self.text_3.pack()

        # start button
        self.but_start = Button(parent, text="Выполнить",command=self.start)
        self.but_start.pack()


if __name__ == "__main__":
        root = Tk()
        root.minsize(width=500, height=400)
        gui = Gui(root)
        root.mainloop()
